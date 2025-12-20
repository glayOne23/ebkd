from json import dumps

import openpyxl
from apps.main.forms.asesor import AsesorExcelForm, AsesorForm
from apps.main.models import (Asesor, JabatanFungsional, JenjangPendidikan,
                              RumpunIlmu)
from apps.services.apigateway import apigateway
from apps.services.utils import profilesync
from django.contrib import messages
from django.contrib.auth.mixins import AccessMixin, LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Count, Q
from decimal import Decimal

from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.safestring import mark_safe
from django.views import View
from django.views.generic import (CreateView, DeleteView, ListView,
                                  TemplateView, UpdateView)


# =====================================================================================================
#                                               MIXINS
# =====================================================================================================
def in_grup(user, grup_name):
    if user:
        return user.groups.filter(name=grup_name).count() > 0
    return False


class AdminRequiredMixin(AccessMixin):
    """Mixin untuk membatasi akses hanya untuk grup personalia."""
    def dispatch(self, request, *args, **kwargs):
        if not in_grup(request.user, 'admin'):
            return HttpResponseForbidden("Anda tidak memiliki hak akses.")
        return super().dispatch(request, *args, **kwargs)


class CustomTemplateBaseMixin(LoginRequiredMixin):
    """Mixin dasar untuk semua view"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # ===[Select CSS and JS Files]===
        context['datatables']       = True
        context['select2']          = True
        context['summernote']       = False
        context['maxlength']        = False
        context['inputmask']        = False
        context['duallistbox']      = False
        context['moment']           = False
        context['daterange']        = False
        context['chartjs']          = False
        return context


# =====================================================================================================
#                                               LOAD PAGE
# =====================================================================================================
class AdminAsesorListView(AdminRequiredMixin, CustomTemplateBaseMixin, ListView):
    model = Asesor
    template_name = 'main/admin/asesor/table.html'
    context_object_name = 'data_asesor'

    def get_queryset(self):
        return self.model.objects.order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['asesor_excel_form'] = AsesorExcelForm
        return context


class AdminAsesorCreateView(AdminRequiredMixin, CustomTemplateBaseMixin, CreateView):
    model = Asesor
    template_name = 'main/admin/asesor/add.html'
    form_class = AsesorForm
    success_url = reverse_lazy('main:admin.asesor.table')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Data tipe AIK berhasil dibuat!")
        return response


class AdminAsesorUpdateView(AdminRequiredMixin, CustomTemplateBaseMixin, UpdateView):
    model = Asesor
    template_name = 'main/admin/asesor/add.html'
    form_class = AsesorForm
    success_url = reverse_lazy('main:admin.asesor.table')

    def get_object(self, queryset=None):
        return get_object_or_404(self.model, id=self.kwargs.get('id'))

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Data tipe AIK berhasil diperbarui")
        return response


# =====================================================================================================
#                                                SERVICE
# =====================================================================================================
class AdminAsesorDeleteListView(AdminRequiredMixin, View):
    def post(self, request):
        list_id = request.POST.getlist('list_id')

        if not list_id:
            messages.error(request, 'Invalid request!')
            return redirect('main:admin.asesor.table')

        try:
            with transaction.atomic():
                # delete sekaligus (lebih cepat & aman)
                deleted_count, _ = Asesor.objects.filter(id__in=list_id).delete()

            if deleted_count > 0:
                messages.success(request, f'{deleted_count} data berhasil dihapus')
            else:
                messages.warning(request, 'Tidak ada data dipilih untuk dihapus')

        except Exception as e:
            messages.error(request, f'Terdapat kesalahan ketika menghapus data {str(e)}')

        return redirect('main:admin.asesor.table')



class AdminAsesorExcelImportView(AdminRequiredMixin, View):

    def get_user_by_nidn(self, namauser, nidn):
        user = User.objects.select_related("profile").filter(profile__nidn=nidn).first()

        if not user:
            raise Exception(f"Asesor tidak ditemukan (NAMA='{namauser}', NIDN='{nidn}')" )

        return user


    def parse_rumpun(self, value, tipe):
        """
        value: 'A01 - Ilmu Sosial'
        tipe: rumpun | subrumpun | bidang
        """
        if not value:
            return None

        try:
            kode, nama = value.split('-', 1)
            obj, _ = RumpunIlmu.objects.update_or_create(
                kode_rumpun=kode.strip(),
                defaults={
                    'nama': nama.strip(),
                    'tipe_rumpun': tipe
                }
            )
            return obj
        except Exception:
            raise ValueError(f"Format {tipe} tidak valid: '{value}'")


    def _import_regular(self, workbook):
        user_cache = {}
        errors = []   # ⬅️ kumpulkan error

        for sheet in workbook.worksheets:

            if sheet.max_row < 3:
                continue

            for idx, row in enumerate(
                sheet.iter_rows(min_row=3, values_only=True),
                start=3
            ):
                if not any(row):
                    continue

                try:
                    namauser = str(row[1]).strip()

                    raw_nira = row[2]
                    if raw_nira is None:
                        nira = None
                    elif isinstance(raw_nira, float):
                        nira = format(Decimal(str(raw_nira)), 'f')
                    else:
                        nira = str(raw_nira).strip()

                    nidn = str(row[3]).strip()

                    jafung = (
                        JabatanFungsional.objects.get(nama=str(row[4]).strip())
                        if row[4] else None
                    )

                    pendidikanterakhir = (
                        JenjangPendidikan.objects.get(nama=str(row[5]).strip())
                        if row[5] else None
                    )

                    statusasesor = str(row[7]).strip() == 'Aktif'

                    rumpunilmu = self.parse_rumpun(row[8], 'rumpun')
                    subrumpunilmu = self.parse_rumpun(row[9], 'subrumpun')
                    bidangilmu = self.parse_rumpun(row[10], 'bidang')

                    cache_key = f"{nidn}"

                    if cache_key in user_cache:
                        peserta = user_cache[cache_key]
                    else:
                        peserta = self.get_user_by_nidn(namauser, nidn)
                        user_cache[cache_key] = peserta

                    Asesor.objects.update_or_create(
                        user=peserta,
                        defaults={
                            'nira': nira,
                            'jabatanfungsional': jafung,
                            'pendidikanterakhir': pendidikanterakhir,
                            'rumpunilmu': rumpunilmu,
                            'subrumpunilmu': subrumpunilmu,
                            'bidangilmu': bidangilmu,
                            'aktif': statusasesor,
                        }
                    )

                except Exception as e:
                    errors.append(
                        f"Sheet '{sheet.title}' baris {idx}: {e}"
                    )

        return errors


    def post(self, request, *args, **kwargs):
        form = AsesorExcelForm(request.POST, request.FILES)

        if not form.is_valid():
            messages.error(request, f"Form tidak valid. {form.get_form_errors()}")
            return redirect('main:admin.asesor.table')

        try:
            workbook = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
        except Exception as e:
            messages.error(request, f"Gagal membaca file Excel: {e}")
            return redirect('main:admin.asesor.table')

        try:
            with transaction.atomic():
                errors = self._import_regular(workbook)

                if errors:
                    error_text = "<br>".join(errors)
                    messages.warning(request, mark_safe(f"Import selesai dengan error:<br>{error_text}"))
                    return redirect('main:admin.asesor.table')

        except Exception as e:
            messages.error(request, f"Gagal impor Excel. {e}")
            return redirect('main:admin.asesor.table')

        messages.success(
            request,
            "Import selesai. Semua data berhasil diunggah."
        )
        return redirect('main:admin.asesor.table')
